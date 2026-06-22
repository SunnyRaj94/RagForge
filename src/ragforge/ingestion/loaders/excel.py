from __future__ import annotations

import os
from pathlib import Path
from typing import Any

import openpyxl

from ragforge.config import get_config_path
from ragforge.ingestion.loaders.base import BaseDocumentLoader, register_loader
from ragforge.ingestion.models import DocumentUnit, LoadedBlock
from ragforge.ingestion.processors import BaseUnitProcessor
from ragforge.ingestion.utils import hash_file, normalize_whitespace


@register_loader(".xlsx", ".xls")
class ExcelLoader(BaseUnitProcessor, BaseDocumentLoader):
    def _load_loader_config(self, config_path: str | None) -> dict[str, Any]:
        if config_path is None:
            config_path = get_config_path()
        if not os.path.exists(config_path):
            return {}
        import yaml

        with open(config_path, "r", encoding="utf-8") as fh:
            return yaml.safe_load(fh) or {}

    def _build_units(
        self, file_path: str, config_path: str | None = None
    ) -> list[DocumentUnit]:
        config = self._load_loader_config(config_path)
        excel_config = config.get("excel", {})
        row_batch_size = excel_config.get("row_batch_size", 15)
        header_row_idx = excel_config.get("header_row", 1) - 1
        sheet_mode = excel_config.get("sheet_mode", "all")

        wb = openpyxl.load_workbook(file_path, data_only=True)
        units: list[DocumentUnit] = []
        file_hash = hash_file(file_path)
        sheets_to_parse = wb.sheetnames
        if sheet_mode == "first" and sheets_to_parse:
            sheets_to_parse = [sheets_to_parse[0]]

        unit_num = 1
        for sheet_name in sheets_to_parse:
            sheet = wb[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue

            headers = []
            if header_row_idx < len(rows):
                headers = [
                    str(cell) if cell is not None else ""
                    for cell in rows[header_row_idx]
                ]

            data_rows = rows[header_row_idx + 1 :]
            if not data_rows:
                continue

            for start in range(0, len(data_rows), row_batch_size):
                batch = data_rows[start : start + row_batch_size]
                batch_lines = []
                if headers:
                    batch_lines.append("| " + " | ".join(headers) + " |")
                    batch_lines.append("| " + " | ".join("---" for _ in headers) + " |")
                for row in batch:
                    row_cells = [str(cell) if cell is not None else "" for cell in row]
                    if len(row_cells) < len(headers):
                        row_cells += [""] * (len(headers) - len(row_cells))
                    elif len(row_cells) > len(headers) and headers:
                        row_cells = row_cells[: len(headers)]
                    batch_lines.append("| " + " | ".join(row_cells) + " |")

                text = normalize_whitespace(
                    f"Sheet: {sheet_name}\n" + "\n".join(batch_lines)
                )
                if not text:
                    continue
                units.append(
                    DocumentUnit(
                        doc_id=file_hash,
                        source_path=file_path,
                        source_name=Path(file_path).name,
                        file_type="excel",
                        unit_num=unit_num,
                        raw_text=text,
                        metadata={
                            "source": file_path,
                            "filename": Path(file_path).name,
                            "file_type": "excel",
                            "sheet": sheet_name,
                            "start_row": start + header_row_idx + 2,
                            "end_row": start + header_row_idx + 1 + len(batch),
                        },
                        unit_kind="sheet_window",
                    )
                )
                unit_num += 1

        return units

    def load(self, file_path: str, config_path: str | None = None) -> list[LoadedBlock]:
        units = self._build_units(file_path, config_path=config_path)
        return self.process_units(units, config_path=config_path)

    async def load_async(
        self, file_path: str, config_path: str | None = None
    ) -> list[LoadedBlock]:
        units = self._build_units(file_path, config_path=config_path)
        return await self.process_units_async(units, config_path=config_path)
